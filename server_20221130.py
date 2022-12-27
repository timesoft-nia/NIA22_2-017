# -*- coding:utf-8 -*-

import os
import torch

import commons
import utils
from models import SynthesizerTrn
from text.symbols import symbols
from text import text_to_sequence

import server_config_20221130 as cf
from flask import Flask, render_template, request, redirect
import nltk

import soundfile as sf
from operator import itemgetter
import re
from openpyxl import load_workbook


# Flask 기반 웹UI 선언
nltk.download('punkt')
app = Flask(__name__)


# 하이퍼파라미터 로드
hps = utils.get_hparams_from_file("./configs/nia22.json")
net_g = SynthesizerTrn(
    len(symbols),
    hps.data.filter_length // 2 + 1,
    hps.train.segment_size // hps.data.hop_length,
    n_speakers=hps.data.n_speakers,
    **hps.model).cuda()
_ = net_g.eval()

# 모델 로드
_ = utils.load_checkpoint("./logs/vits_nia22.pth", net_g, None)

# 화자정보 로드
read_from = load_workbook('speaker_info_20221130.xlsx')
read_sheet = read_from['Sheet1']
num = 1
for row in range(1, read_sheet.max_row + 1):
    speaker_id_info = str(read_sheet.cell(row=row, column=1).value)
    speaker_age_info = str(read_sheet.cell(row=row, column=2).value)
    speaker_sex_info = str(read_sheet.cell(row=row, column=3).value)
    speaker_job_info = str(read_sheet.cell(row=row, column=4).value)
    speaker_time_info = str(read_sheet.cell(row=row, column=5).value)

    cf.INPUT_FIELDS['pv' + str(num)]['label'] = cf.INPUT_FIELDS['pv' + str(num)]['label'] + ' [' + speaker_age_info + '/' + speaker_sex_info + '/' + speaker_job_info + '/' + speaker_time_info + ']'
    num = num + 1


# 음성 합성 파일 저장 디렉토리 초기화
def clear_storage(ip_addr):
    for f in os.listdir(cf.wave_storage):
        file_split = f.split(ip_addr)
        if len(file_split) != 1:
            os.remove(os.path.join(cf.wave_storage, f))


# 웹 초기화면 및 화자모델 로드 페이지
@app.route('/', methods=['GET'])
def index():
    return render_template('index.html', input_fields=cf.INPUT_FIELDS.values())


# 텍스트 -> 음성 합성 페이지
@app.route('/generate', methods=['GET', 'POST'])
def generate():
    if request.method == 'GET':
        return redirect('/')

    # 음성합성 시 음원저장 디렉토리 초기화
    clear_storage(request.environ.get('HTTP_X_REAL_IP', request.remote_addr))

    form = request.form
    text = form.get('text', '').strip()

    # 모델 정의
    MODELS = {
        "pv1": "3",
        "pv2": "4",
        "pv3": "5",
        "pv4": "6",
        "pv5": "7",
        "pv6": "8",
        "pv7": "9",
        "pv8": "10",
        "pv9": "11",
        "pv10": "12",
        "pv11": "13",
        "pv12": "14",
        "pv13": "15",
        "pv14": "16",
        "pv15": "17",
        "pv16": "19",
        "pv17": "20",
        "pv18": "21",
        "pv19": "22",
        "pv20": "23",
        "pv21": "24",
        "pv22": "25",
        "pv23": "26",
        "pv24": "27",
        "pv25": "28",
        "pv26": "29",
        "pv27": "30",
        "pv28": "31",
        "pv29": "32",
        "pv30": "33",
        "pv31": "34",
        "pv32": "35",
        "pv33": "36",
        "pv34": "37",
        "pv35": "38",
        "pv36": "39",
        "pv37": "40",
        "pv38": "41",
        "pv39": "42",
        "pv40": "43",
        "pv41": "44",
        "pv42": "45",
        "pv43": "46",
        "pv44": "47",
        "pv45": "48",
        "pv46": "49",
        "pv47": "50",
        "pv48": "51",
        "pv49": "52",
        "pv50": "53",
        "pv51": "54",
        "pv52": "55",
        "pv53": "56",
        "pv54": "57",
        "pv55": "58",
        "pv56": "59",
        "pv57": "60",
        "pv58": "61",
        "pv59": "62",
        "pv60": "63",
        "pv61": "64",
        "pv62": "65",
        "pv63": "66",
        "pv64": "67",
        "pv65": "68",
        "pv66": "69",
        "pv67": "70",
        "pv68": "71",
        "pv69": "72",
        "pv70": "73",
        "pv71": "74",
        "pv72": "75",
        "pv73": "76",
        "pv74": "77",
        "pv75": "78",
        "pv76": "79",
        "pv77": "80",
        "pv78": "81",
        "pv79": "82",
        "pv80": "83",
        "pv81": "84",
        "pv82": "85",
        "pv83": "86",
        "pv84": "87",
        "pv85": "88",
        "pv86": "89"
    }

    # 해당 화자모델로 텍스트 -> 음성 합성
    audio = []

    # 텍스트 전처리
    if text != '':
        pattern = r'\([^)]*\)'
        text = re.sub(pattern=pattern, repl='', string=text)
        text = text.strip()
        text = text.replace('!', '.')

        if text[-1] != '.':
            text = text + '.'

        # 소수점 표현일 경우 처리
        text_split = text.split('.')
        for text_num in range(len(text_split) - 2):
            text_space_split = text_split[text_num].strip()
            text_space_next_split = text_split[text_num + 1].strip()

            text_space_split_result = text_space_split[-1].isdigit()

            if text_space_split_result is True:
                text_split[text_num] = text_split[text_num] + '.' + text_space_next_split
                del text_split[text_num + 1]

        for text_num in range(len(text_split)):
            print(text_split[text_num])
            if text_split[text_num].strip() == '':
                continue
            for k in MODELS.keys():
                if form.get(k, None) is not None:
                    stn_tst = get_text(text_split[text_num].strip())
                    with torch.no_grad():
                        x_tst = stn_tst.cuda().unsqueeze(0)
                        x_tst_lengths = torch.LongTensor([stn_tst.size(0)]).cuda()
                        sid = torch.LongTensor([cf.INPUT_FIELDS[k]['speaker_id']]).cuda()
                        inference_audio = \
                            net_g.infer(x_tst, x_tst_lengths, sid=sid, noise_scale=.667, noise_scale_w=0.8,
                                        length_scale=1)[
                                0][0, 0].data.cpu().float().numpy()

                    file_name = './static/audio/nia22_speaker_' + str(cf.INPUT_FIELDS[k]['speaker_id']) + '_' + str(
                        text_num + 1) + '_' + request.environ.get('HTTP_X_REAL_IP', request.remote_addr) + '.wav'

                    sf.write(file_name, inference_audio, hps.data.sampling_rate)

                    audio.append({
                        'label': cf.INPUT_FIELDS[k]['label'],
                        'src': file_name,
                        'text': str(text_split[text_num]).strip()
                    })

    # 화자수 조회
    audio = sorted(audio, key=itemgetter('label'))

    audio_length = len(audio)
    audio_list_num = []
    audio_dup_remove = []

    for i in range(audio_length):
        if audio[i]['label'] not in audio_dup_remove:
            audio_dup_remove.append(audio[i]['label'])

    try:
        audio_speaker_num = len(audio_dup_remove)
        audio_real_length = audio_length / audio_speaker_num
        for j in range(0, audio_length, int(audio_real_length)):
            try:
                audio_list_num.append(j)
            except ZeroDivisionError as e:
                continue
    except ZeroDivisionError as e:
        pass
    except ValueError as e:
        pass

    # 음성 합성 파일 정보를 웹으로 반환
    return render_template('generate.html', input_fields=cf.INPUT_FIELDS.values(), audio=audio,
                           audio_list_num=audio_list_num)


# 사용자 텍스트 변환
def get_text(text):
    text_norm = text_to_sequence(text, hps.data.text_cleaners)
    if hps.data.add_blank:
        text_norm = commons.intersperse(text_norm, 0)
    text_norm = torch.LongTensor(text_norm)
    return text_norm


if __name__ == '__main__':
    # 웹 IP 및 포트 정의
    app.run(host='0.0.0.0', port=5002, debug=True, threaded=True)
